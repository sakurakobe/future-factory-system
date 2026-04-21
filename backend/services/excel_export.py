"""
================================================================================
Excel 报告导出服务
================================================================================
功能：
    将评估结果导出为 Excel 文件，包含所有题目的评估详情。

导出列：
    - 方面、子类、细项（分类层级）
    - 评测题目
    - 当前等级（沟通等级）
    - 当前得分
    - 目标等级
    - 目标分值
    - 企业现状
    - 沟通内容
    - 改进建议

使用方式：
    from services.excel_export import export_excel
    filepath = export_excel(project_id, db)

依赖：
    openpyxl 库，用于 Excel 文件的创建
================================================================================
"""
import json
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from models.category import MajorCategory, SubCategory, SubSubCategory
from models.question import Question
from models.answer import AssessmentAnswer
from models.project import AssessmentProject


def _should_show_question(question, company_type: str) -> bool:
    """判断题目是否应该显示（与 categories API 保持一致）"""
    if company_type == "通用":
        return True
    title = question.title
    if company_type == "离散":
        return "流程行业：" not in title
    if company_type == "流程":
        return "离散行业：" not in title
    return True


def _generate_suggestion(answer, question) -> str:
    """根据当前等级和目标等级生成改进建议"""
    LEVEL_ORDER = ['A', 'B', 'C', 'D', 'E', 'F']
    if not answer.selected_level or not answer.target_level:
        return ''
    if answer.selected_level == answer.target_level:
        return '已达目标，持续优化。'

    current_idx = LEVEL_ORDER.index(answer.selected_level) if answer.selected_level in LEVEL_ORDER else 0
    target_idx = LEVEL_ORDER.index(answer.target_level) if answer.target_level in LEVEL_ORDER else 5

    options = json.loads(question.options_json)
    suggestions = []
    for i in range(current_idx + 1, target_idx + 1):
        for opt in options:
            if opt['level'] == LEVEL_ORDER[i] and opt.get('description'):
                desc = opt['description'][:80]
                suggestions.append(desc)
                break

    return '、'.join(suggestions) if suggestions else ''


def export_excel(project_id: int, db: Session) -> str:
    """
    导出项目的评估结果为 Excel 文件。

    参数：
        project_id: 项目ID
        db: 数据库会话

    返回：
        生成的 Excel 文件路径
    """
    # 获取项目信息
    project = db.query(AssessmentProject).filter(
        AssessmentProject.id == project_id
    ).first()
    if not project:
        raise ValueError(f"项目 {project_id} 不存在")

    company_type = project.company_type or "通用"

    # 创建 Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "诊断评估结果"

    # 样式定义
    header_font = Font(name='宋体', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    data_font = Font(name='宋体', size=10)
    wrap_alignment = Alignment(wrap_text=True, vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 表头
    headers = [
        '方面', '子类', '细项', '评测题目',
        '当前等级', '当前得分', '满分',
        '目标等级', '目标分值',
        '企业现状', '沟通内容', '改进建议'
    ]

    # 写入表头
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # 列宽设置
    col_widths = [20, 20, 20, 40, 10, 10, 10, 10, 10, 30, 30, 40]
    for idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    # 数据行
    row_num = 2
    majors = db.query(MajorCategory).order_by(MajorCategory.sort_order).all()

    for major in majors:
        subs = db.query(SubCategory).filter(
            SubCategory.major_category_id == major.id
        ).order_by(SubCategory.sort_order).all()

        for sub in subs:
            subsubs = db.query(SubSubCategory).filter(
                SubSubCategory.sub_category_id == sub.id
            ).order_by(SubSubCategory.sort_order).all()

            for ss in subsubs:
                questions = db.query(Question).filter(
                    Question.sub_sub_category_id == ss.id
                ).order_by(Question.sort_order).all()

                for q in questions:
                    # 根据 company_type 筛选题目
                    if not _should_show_question(q, company_type):
                        continue

                    # 获取回答
                    answer = db.query(AssessmentAnswer).filter(
                        AssessmentAnswer.project_id == project_id,
                        AssessmentAnswer.question_id == q.id
                    ).first()

                    # 当前等级和得分
                    current_level = answer.selected_level if answer and answer.selected_level else ''
                    current_score = answer.score if answer and answer.score else 0

                    # 目标等级和目标分值
                    target_level = answer.target_level if answer and answer.target_level else ''
                    target_score = 0
                    if answer and answer.target_level:
                        options = json.loads(q.options_json)
                        for opt in options:
                            if opt['level'] == answer.target_level:
                                target_score = opt['score']
                                break

                    # 企业现状和沟通内容
                    company_status = answer.company_status if answer and answer.company_status else ''
                    comm_content = answer.communication_content if answer and answer.communication_content else ''

                    # 改进建议
                    suggestion = _generate_suggestion(answer, q) if answer else ''

                    # 写入数据行
                    row_data = [
                        major.name, sub.name, ss.name, q.title,
                        current_level, current_score, q.max_score,
                        target_level, target_score,
                        company_status, comm_content, suggestion
                    ]

                    for col_idx, value in enumerate(row_data, 1):
                        cell = ws.cell(row=row_num, column=col_idx, value=value)
                        cell.font = data_font
                        cell.alignment = wrap_alignment
                        cell.border = thin_border

                    row_num += 1

    # 设置打印区域和页面布局
    ws.freeze_panes = 'A2'  # 冻结表头
    ws.auto_filter.ref = ws.dimensions

    # 保存文件
    import os
    from config import UPLOAD_DIR
    filename = f"评估结果_{project.company_name}.xlsx"
    filepath = os.path.join(UPLOAD_DIR, filename)
    wb.save(filepath)

    return filepath, filename
