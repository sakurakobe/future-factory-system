"""
================================================================================
XLSX 题库导入服务
================================================================================
功能：
    将 XLSX 格式的题库文件导入到 SQLite 数据库中。
    导入的数据包含完整的分类层级和题目等级选项。

XLSX 文件结构（未来工厂标准诊断评估题库.xlsx）：
    A列 - 主要方面（如 "1.总体架构"）
    B列 - 子类（如 "1.技术支撑"）
    C列 - 二级分类/细项（如 "1.新一代信息技术"）
    D列 - 等级标签D（如 "通用"）
    E列 - 等级标签E（如 "数字化车间"）
    F列 - 等级标签F（如 "智能工厂"）
    G列 - 等级标签G（如 "未来工厂"）
    H列 - 评估题目文字（重要：仅在新题目的第一行出现！同一题目的后续行H列为空）
    I列 - 等级标识（A/B/C/D/E/F）
    J列 - 等级描述文字
    K列 - 该等级对应的分值
    L列 - 应达等级（目标等级）

关键理解：
    H列（题目文字）只在每个新题目的第一行出现，后续行 H 列为 None。
    这意味着：
    - 当 H 列有新值 → 创建新题目，并开始添加第一个等级选项
    - 当 H 列为空 → 当前行是上一题目的一个等级选项（I列有A-F标识）
    每个题目通常有 6 个等级选项（A到F），分布在连续的 6 行中。

导入流程：
    1. 清空已有分类和题目数据
    2. 逐行解析 XLSX，识别分类层级和题目
    3. 将等级选项（A-F）以 JSON 格式存入题目表的 options_json 字段
    4. 自动计算各级分类的 max_score（题目分值之和）

================================================================================
"""
import json
import openpyxl
from sqlalchemy.orm import Session
from models.category import MajorCategory, SubCategory, SubSubCategory
from models.question import Question
from config import XLSX_PATH


def import_xlsx(db: Session):
    """
    从 XLSX 文件导入完整的题库数据到数据库。

    执行流程：
    1. 清空所有已有数据（防止重复导入）
    2. 逐行解析 XLSX 文件，识别分类层级
    3. 收集题目及其等级选项（A-F）
    4. 批量写入数据库
    5. 重新计算各级分类的 max_score

    注意：此函数会清空所有分类和题目数据，请谨慎调用！
    """
    # 第一步：清空所有已有数据（从子到父，避免外键约束问题）
    db.query(Question).delete()
    db.query(SubSubCategory).delete()
    db.query(SubCategory).delete()
    db.query(MajorCategory).delete()
    db.commit()

    # 第二步：打开 XLSX 文件，读取第一个工作表
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb[wb.sheetnames[0]]

    # 当前遍历到的分类对象（随行数据动态更新）
    current_major = None
    current_sub = None
    current_subsub = None

    # 排序序号计数器
    major_order = 0
    sub_order = 0
    subsub_order = 0
    question_order = 0

    # 使用列表暂存题目数据，最后统一写入数据库
    # 每个元素格式：{"title": "...", "sub_sub_category_id": N, "sort_order": N, "options": [...]}
    pending_questions = []
    current_q = None  # 当前正在处理的题目（用于持续添加等级选项）

    # 等级标签（D/E/F/G列的值，用于报告的等级说明）
    level_label_d = level_label_e = level_label_f = level_label_g = None

    # 逐行解析（从第2行开始，跳过表头）
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), start=2):
        major_val = row[0]       # A列: 主要方面
        sub_val = row[1]         # B列: 子类
        subsub_val = row[2]      # C列: 二级分类/细项
        level_label_d = row[3]   # D列: 等级标签D
        level_label_e = row[4]   # E列: 等级标签E
        level_label_f = row[5]   # F列: 等级标签F
        level_label_g = row[6]   # G列: 等级标签G
        question_title = row[7]  # H列: 评估题目（仅在新题目首行出现）
        level = row[8]           # I列: A/B/C/D/E/F 等级标识
        level_desc = row[9]      # J列: 等级描述
        score = row[10]          # K列: 对应分值
        target_level = row[11]   # L列: 应达等级

        # ---- 识别并创建主要方面（如"赋能保障"） ----
        if major_val and str(major_val).strip():
            major_name = str(major_val).strip()
            current_major = MajorCategory(name=major_name, sort_order=major_order + 1, max_score=0)
            major_order += 1
            db.add(current_major)
            db.flush()  # 立即获取自增ID，不提交事务
            sub_order = 0
            subsub_order = 0

        # ---- 识别并创建子类（如"技术支撑"） ----
        if sub_val and str(sub_val).strip():
            sub_name = str(sub_val).strip()
            current_sub = SubCategory(
                major_category_id=current_major.id,
                name=sub_name,
                sort_order=sub_order + 1,
                max_score=0
            )
            db.add(current_sub)
            db.flush()
            sub_order += 1
            subsub_order = 0

        # ---- 识别并创建细项（如"新一代信息技术"） ----
        if subsub_val and str(subsub_val).strip():
            subsub_name = str(subsub_val).strip()
            current_subsub = SubSubCategory(
                sub_category_id=current_sub.id,
                name=subsub_name,
                sort_order=subsub_order + 1,
                max_score=0
            )
            db.add(current_subsub)
            db.flush()
            subsub_order += 1

        # ---- 处理题目 ----
        # 关键逻辑：H列有值 = 新题目开始；H列为空 = 当前题目的等级选项
        if question_title and str(question_title).strip():
            # 新题目：创建题目对象，准备接收等级选项
            q_title = str(question_title).strip()
            question_order += 1
            current_q = {
                "title": q_title,
                "sub_sub_category_id": current_subsub.id,
                "sort_order": question_order,
                "options": [],  # 等级选项列表，初始为空
            }
            pending_questions.append(current_q)

        # ---- 添加等级选项到当前题目 ----
        if current_q and level and str(level).strip():
            current_q["options"].append({
                "level": str(level).strip(),          # 如 "A", "B", ...
                "score": float(score) if score else 0,
                "description": str(level_desc) if level_desc else "",
                "target_level": str(target_level) if target_level else "",
            })

    # ---- 保存等级标签映射 ----
    level_labels = {
        "D": str(level_label_d) if level_label_d else "通用",
        "E": str(level_label_e) if level_label_e else "数字化车间",
        "F": str(level_label_f) if level_label_f else "智能工厂",
        "G": str(level_label_g) if level_label_g else "未来工厂",
    }

    # 第三步：将题目写入数据库
    for q_data in pending_questions:
        # max_score: 多选题取所有选项分值之和（用户可选择多项），单选题取最高等级分值
        is_multi = "（多选题）" in q_data["title"] or "多选题" in q_data["title"]
        max_score = sum((o["score"] for o in q_data["options"])) if is_multi else max((o["score"] for o in q_data["options"]), default=0)

        # 从题目标目中识别行业类型
        title = q_data["title"]
        if "离散行业" in title and "离散行业：" not in title:
            # 兼容：标题中含"离散行业"但不含冒号（全角/半角都可能）
            industry_type = "离散"
        elif "流程行业" in title and "流程行业：" not in title:
            industry_type = "流程"
        elif "离散行业：" in title or "离散行业" == title.split("：")[0].split(":")[0]:
            industry_type = "离散"
        elif "流程行业：" in title or "流程行业" == title.split("：")[0].split(":")[0]:
            industry_type = "流程"
        else:
            industry_type = "通用"

        q = Question(
            sub_sub_category_id=q_data["sub_sub_category_id"],
            sort_order=q_data["sort_order"],
            title=q_data["title"],
            max_score=max_score,
            industry_type=industry_type,
            level_labels=json.dumps(level_labels, ensure_ascii=False),
            options_json=json.dumps(q_data["options"], ensure_ascii=False),
        )
        db.add(q)

    db.commit()

    # 第四步：重新计算各级分类的 max_score
    _recalc_scores(db)


def _recalc_scores(db: Session):
    """
    自底向上重新计算各分类的 max_score。

    计算顺序：细项 → 子类 → 主要方面
    每个分类的 max_score = 其下所有题目 max_score 之和
    """
    # 计算细项 max_score = 其下所有题目 max_score 之和
    subsubs = db.query(SubSubCategory).all()
    for ss in subsubs:
        questions = db.query(Question).filter(Question.sub_sub_category_id == ss.id).all()
        ss.max_score = sum(q.max_score for q in questions)

    # 计算子类 max_score = 其下所有细项 max_score 之和
    subs = db.query(SubCategory).all()
    for s in subs:
        subsubs = db.query(SubSubCategory).filter(SubSubCategory.sub_category_id == s.id).all()
        s.max_score = sum(ss.max_score for ss in subsubs)

    # 计算主要方面 max_score = 其下所有子类 max_score 之和
    majors = db.query(MajorCategory).all()
    for m in majors:
        subs = db.query(SubCategory).filter(SubCategory.major_category_id == m.id).all()
        m.max_score = sum(s.max_score for s in subs)

    db.commit()
